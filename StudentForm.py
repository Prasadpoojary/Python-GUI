from tkinter.ttk import*
from openpyxl import *
from tkinter import*
from tkinter import filedialog
from tkinter import messagebox

try:
   wb=load_workbook("C:\\Users\\User\\Documents\\college.xlsx")
   sheet=wb.active
except FileNotFoundError:
   messagebox.showwarning("Warning","Please create a 'college.xlsx' file in 'C:/Users/User/Documents'")
def settings():
    sheet.column_dimensions['A'].width=25
    sheet.column_dimensions['B'].width=12
    sheet.column_dimensions['C'].width=25
    sheet.column_dimensions['D'].width=25
    sheet.column_dimensions['E'].width=20
    sheet.column_dimensions['F'].width=14
    sheet.column_dimensions['G'].width=30
    sheet.column_dimensions['H'].width=15
    sheet.column_dimensions['I'].width=15
    sheet.column_dimensions['J'].width=25
    sheet.column_dimensions['K'].width=13
    sheet.column_dimensions['L'].width=25
    sheet.column_dimensions['M'].width=15
    sheet.column_dimensions['N'].width=15
    sheet.column_dimensions['O'].width=20
    sheet.column_dimensions['P'].width=18
    sheet.column_dimensions['Q'].width=20
    sheet.column_dimensions['R'].width=20
    sheet.cell(row=1,column=1).value="NAME"
    sheet.cell(row=1,column=2).value="SEX"
    sheet.cell(row=1,column=3).value="DATE OF BIRTH"
    sheet.cell(row=1,column=4).value="FATHER NAME"
    sheet.cell(row=1,column=5).value="AADHAR"
    sheet.cell(row=1,column=6).value="MOBILE"
    sheet.cell(row=1,column=7).value="E MAIL"
    sheet.cell(row=1,column=8).value="RELIGION"
    sheet.cell(row=1,column=9).value="STATUS"
    sheet.cell(row=1,column=10).value="INCOME CERT"
    sheet.cell(row=1,column=11).value="INCOME"
    sheet.cell(row=1,column=12).value="CASTE CERT"
    sheet.cell(row=1,column=13).value="CATOGORY"
    sheet.cell(row=1,column=14).value="CASTE"
    sheet.cell(row=1,column=15).value="BANK"
    sheet.cell(row=1,column=16).value="IFSC"
    sheet.cell(row=1,column=17).value="BRANCH"
    sheet.cell(row=1,column=18).value="ACCOUNT NUMER"

def clear():
    ent1.delete(0,END)
    ent2.delete(0,END)
    ent3.delete(0,END)
    ent4.delete(0,END)
    ent5.delete(0,END)
    ent6.delete(0,END)
    ent7.delete(0,END)
    ent8.delete(0,END)
    ent9.delete(0,END)
    ent10.delete(0,END)
    ent11.delete(0,END)
    ent12.delete(0,END)
    c1.delete(0,END)
    c2.delete(0,END)
    c3.delete(0,END)
    c4.delete(0,END)
    c5.delete(0,END)
    c6.delete(0,END)
    c7.delete(0,END)
    c1.set("      select")
    c2.set("DD")
    c3.set("Month")
    c4.set("YYYY ")
    c5.set("      select")
    c6.set("      select")
    c7.set("      select")


def fent1(event):
    c2.focus_set()
def fent2(event):
    c3.focus_set()
def fent3(event):
    c4.focus_set()
def fent4(event):
    ent3.focus_set()
def fent5(event):
    ent4.focus_set()
def fent6(event):
    c1.focus_set()
def fent7(event):
    ent2.focus_set()
def fent8(event):
    c5.focus_set()
def fent9(event):
    ent5.focus_set()
def fent10(event):
    ent6.focus_set()
def fent11(event):
    ent7.focus_set()
def fent12(event):
    c6.focus_set()
def fent13(event):
    c7.focus_set()
def fent14(event):
    ent8.focus_set()
def fent15(event):
    ent9.focus_set()
def fent16(event):
    ent10.focus_set()
def fent17(event):
    ent11.focus_set()
def fent18(event):
    ent12.focus_set()


  
def insert():
    if  ent1.get()=="" or ent2.get()=="" or ent3.get()=="" or ent4.get()=="" or ent5.get()=="" or ent6.get()=="" or ent7.get()=="" or ent8.get()=="" or ent9.get()=="" or ent10.get()=="" or ent11.get()=="" or ent12.get()=="" or c1.get()=="      select" or c2.get()=="DD" or c3.get()=="Month" or c4.get()=="YYYY " or c5.get()=="      select" or c6.get()=="      select" or c7.get()=="      select":
        messagebox.showerror("Error","Please fill every details with valid informations, Do not leave blank Entries")
    elif not ent2.get().endswith("@gmail.com"):
            messagebox.showwarning("Warning","Please Enter a valid Email address")
                             
    else:
        settings()
        mrow=sheet.max_row
        mcol=sheet.max_column
        sheet.cell(row=mrow+1,column=1).value=ent1.get()
        if i.get()==2:
                gen="Female"

        else:
                gen="Male"

        sheet.cell(row=mrow+1,column=2).value=gen
        sheet.cell(row=mrow+1,column=3).value=str(c2.get())+" "+str(c3.get())+" "+str(c4.get())
        sheet.cell(row=mrow+1,column=4).value=ent8.get()
        sheet.cell(row=mrow+1,column=5).value=ent3.get()
        sheet.cell(row=mrow+1,column=6).value=ent4.get()
        sheet.cell(row=mrow+1,column=7).value=ent2.get()
        sheet.cell(row=mrow+1,column=8).value=c1.get()
        sheet.cell(row=mrow+1,column=9).value=c5.get()
        sheet.cell(row=mrow+1,column=10).value=ent5.get()
        sheet.cell(row=mrow+1,column=11).value=ent6.get()
        sheet.cell(row=mrow+1,column=12).value=ent7.get()
        sheet.cell(row=mrow+1,column=13).value=c6.get()
        sheet.cell(row=mrow+1,column=14).value=c7.get()
        sheet.cell(row=mrow+1,column=15).value=ent9.get()
        sheet.cell(row=mrow+1,column=16).value=ent10.get()
        sheet.cell(row=mrow+1,column=17).value=ent11.get()
        sheet.cell(row=mrow+1,column=18).value=ent12.get()
        wb.save("C:\\Users\\User\\Desktop\\college.xlsx")
        clear()
        ent1.focus_set()

if __name__=="__main__":
    
    window=Tk()
    window.geometry("1365x800+0+0")
    window.title("Student Information Form")
    window.configure(background="light grey")
    lab1=Label(window,text="Government First Grade College Tenkanidiyoor, Udupi.",font="arial 24 bold underline",bg="light grey",fg="black")
    lab1.place(x=250,y=10)
    lab2=Label(window,text="Name:",font="none 17 bold",bg="light grey",fg="black")
    ent1=Entry(window,width=20,font="none 17 bold")
    lab2.place(x=20,y=70)
    ent1.place(x=100,y=70)
    ent1.focus_set()
    lab3=Label(window,text="Gender:",font="none 17 bold",bg="light grey",fg="black")
    lab3.place(x=500,y=70)
    i=IntVar()
    radgen1=Radiobutton(window,text="Male",font="none 17",bg="light grey",fg="black",value=1,variable=i)
    radgen1.place(x=600,y=70)
    radgen2=Radiobutton(window,text="Female",font="none 17",bg="light grey",fg="black",value=2,variable=i)
    radgen2.place(x=700,y=70)
    lab4=Label(window,text="Religion:",font="none 17 bold",bg="light grey",fg="black")
    lab4.place(x=1000,y=70)
    c1=Combobox(window,font="none 17",width=10)
    c1["value"]=("Hindu","Muslim","Christian","Sikh","Budd","Jain")
    c1.set("      select")
    c1.place(x=1110,y=70)
    lab5=Label(window,text="Date of Birth:",font="none 17 bold",bg="light grey",fg="black")
    lab5.place(x=20,y=140)
    c2=Combobox(window,font="none 17",width=3)
    c2["value"]=list(range(1,32))
    c2.set("DD")
    c2.place(x=180,y=140)
    lab6=Label(window,text=":",font="none 17 bold",bg="light grey",fg="black")
    lab6.place(x=240,y=140)
    c3=Combobox(window,font="none 17",width=10)
    c3["value"]=("January","February","March","April","May","June","July","August","September","October","November","December")
    c3.set("Month")
    c3.place(x=252,y=140)
    lab7=Label(window,text=":",font="none 17 bold",bg="light grey",fg="black")
    lab7.place(x=404,y=140)
    c4=Combobox(window,font="none 17",width=5)
    c4["value"]=list(range(1990,2019))
    c4.set("YYYY ")
    c4.place(x=416,y=140)
    lab8=Label(window,text="Email:",font="none 17 bold",bg="light grey",fg="black")
    lab8.place(x=700,y=140)
    ent2=Entry(window,width=40,font="none 16")
    ent2.place(x=782,y=142)
    lab9=Label(window,text="Aadhar No. :",font="none 17 bold",bg="light grey",fg="black")
    lab9.place(x=20,y=210)
    ent3=Entry(window,width=20,font="none 16")
    ent3.place(x=170,y=212)
    lab10=Label(window,text="Mobile No. :",font="none 17 bold",bg="light grey",fg="black")
    lab10.place(x=500,y=210)
    ent4=Entry(window,width=14,font="none 16")
    ent4.place(x=640,y=212)
    lab11=Label(window,text="Merital Stutus:",font="none 17 bold",bg="light grey",fg="black")
    lab11.place(x=940,y=210)
    c5=Combobox(window,font="none 17",width=10)
    c5["value"]=("Single","Married","Unmarried")
    c5.set("      select")
    c5.place(x=1110,y=212)
    lab12=Label(window,text="````````````````````````````````````````````````````````````````````````````````````````````````````````````````````````````````````````````````````````````````````````",font="none 17 bold",bg="light grey",fg="blue")
    lab12.place(x=0,y=280)
    lab13=Label(window,text="Income Certificate No.:",font="none 17 bold",bg="light grey",fg="black")
    lab13.place(x=20,y=300)
    ent5=Entry(window,width=20,font="none 16")
    ent5.place(x=280,y=300)
    lab14=Label(window,text="Annual Income:",font="none 17 bold",bg="light grey",fg="black")
    lab14.place(x=20,y=370)
    ent6=Entry(window,width=20,font="none 16")
    ent6.place(x=280,y=370)
    lab15=Label(window,text="Caste Certificate No.:",font="none 17 bold",bg="light grey",fg="black")
    lab15.place(x=20,y=440)
    ent7=Entry(window,width=20,font="none 16")
    ent7.place(x=280,y=440)
    lab16=Label(window,text="Category:",font="none 17 bold",bg="light grey",fg="black")
    lab16.place(x=20,y=510)
    c6=Combobox(window,font="none 17",width=10)
    c6["value"]=("General","OBC 2A","3A","2B","3B","SC","ST")
    c6.set("      select")
    c6.place(x=280,y=510)
    lab17=Label(window,text="Caste:",font="none 17 bold",bg="light grey",fg="black")
    lab17.place(x=20,y=580)
    c7=Combobox(window,font="none 17",width=10)
    c7.set("      select")

    def click():
        if c6.get()== "General":
            c7["value"]=("Konkani","Karaba","Bramhins")
        

        elif c6.get()== "OBC 2A":
            c7["value"]=("Bunts","Billava","Kulal","Eidiga")
        
    btn1=Button(window,text="SUBMIT",font="none 10 bold",command=click)
    btn1.place(x=450,y=512)
    c7.place(x=280,y=580)
    lab18=Label(window,text="Father Name:",bg="light grey",fg="black",font="none 17 bold")
    lab18.place(x=800,y=300)
    ent8=Entry(window,font="none 16",width=20)
    ent8.place(x=955,y=300)
    lab19=Label(window,text="Bank Name:",bg="light grey",fg="black",font="none 17 bold")
    lab19.place(x=800,y=370)
    ent9=Entry(window,font="none 16",width=20)
    ent9.place(x=957,y=370)
    lab20=Label(window,text="IFSC code:",bg="light grey",fg="black",font="none 17 bold")
    lab20.place(x=800,y=440)
    ent10=Entry(window,font="none 16",width=20)
    ent10.place(x=957,y=440)
    lab21=Label(window,text="Branch:",bg="light grey",fg="black",font="none 17 bold")
    lab21.place(x=800,y=510)
    ent11=Entry(window,font="none 16",width=20)
    ent11.place(x=957,y=510)
    lab22=Label(window,text="Account No. :",bg="light grey",fg="black",font="none 17 bold")
    lab22.place(x=800,y=580)
    ent12=Entry(window,font="none 16",width=20)
    ent12.place(x=957,y=580)
    result=IntVar()
    check=Checkbutton(window,text="Fill this informations to Excel Sheet",bg="light grey",font="none 15",variable=result)
    check.place(x=480,y=620)
    def fill():
        if result.get()==0:
            messagebox.showwarning("Warning","Clear all data which you have entered")
            ans=messagebox.askyesno("Clear","Do You want to clear all data?")
            if ans is True:
                clear()
        else:
             insert()
    btn2=Button(window,text="SUBMIT",font="none 18 bold",command=fill)
    btn2.place(x=590,y=652)
    def showmes():
         messagebox.showinfo("info","Please open the excel file on Desktop of name college.xlsx")
         file=messagebox.askyesno("Exit","Do you want to exit from this Page ?")
         if file is True:
             window.destroy()
    btn4=Button(window,text="Exit",font="none 15 bold",width=7,command=showmes)
    btn4.place(x=300,y=652)
    btn5=Button(window,text="Clear",font="none 15 bold",width=7,command=clear)
    btn5.place(x=900,y=652)

    ent1.bind("<Return>",fent1)
    c2.bind("<Return>",fent2) 
    c3.bind("<Return>",fent3)
    c4.bind("<Return>",fent4)
    ent3.bind("<Return>",fent5)
    ent4.bind("<Return>",fent6)
    c1.bind("<Return>",fent7)
    ent2.bind("<Return>",fent8)
    c5.bind("<Return>",fent9)
    ent5.bind("<Return>",fent10)
    ent6.bind("<Return>",fent11)
    ent7.bind("<Return>",fent12)
    c6.bind("<Return>",fent13)
    c7.bind("<Return>",fent14)
    ent8.bind("<Return>",fent15)
    ent9.bind("<Return>",fent16)
    ent10.bind("<Return>",fent17)
    ent11.bind("<Return>",fent18)
    
   
window.mainloop()






 













